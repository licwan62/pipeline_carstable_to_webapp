from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import openpyxl
import yaml

from run_all import scan_cases
from tools.build_publish_site import configure_excel_sources
from tools.build_user_size_templates import configured_stores
from tools.materialize_input_sources import materialize_inputs


def write_config(path: Path, input_dir: Path, middle_dir: Path, output_dir: Path) -> None:
    path.write_text(
        yaml.safe_dump(
            {
                "paths": {
                    "input_dir": str(input_dir),
                    "middle_dir": str(middle_dir),
                    "output_dir": str(output_dir),
                    "logs_dir": str(path.parent / "logs"),
                },
                "file_rules": {"input_patterns": ["*.xlsx", "*.csv"]},
                "input": {
                    "case_name": "combined",
                    "header_aliases": {"S8MAKE": "MAKE"},
                    "match_source": {
                        "name": "{stem}",
                        "label": "{stem}尺码匹配表",
                        "sheet": "{stem}尺码匹配",
                        "header_row": 1,
                        "columns": "MODEL,YEAR,SIZE",
                    },
                },
            },
            allow_unicode=True,
            sort_keys=False,
        ),
        encoding="utf-8",
    )


class InputMaterializationTests(unittest.TestCase):
    def test_multiple_csv_and_excel_files_generate_match_sources_from_filenames(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            input_dir = root / "input"
            input_dir.mkdir()
            (input_dir / "ALL.csv").write_text(
                "MAKE,MODEL,YEAR,确认尺码\nAcura,ADX,2025-2026,YL\n",
                encoding="utf-8-sig",
            )
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "原始数据"
            sheet.append(["MAKE", "MODEL", "YEAR", "确认尺码"])
            sheet.append(["Tesla", "Model Y", "2025", "3XL"])
            workbook.save(input_dir / "TM.xlsx")

            config_path = root / "pipeline.yaml"
            write_config(config_path, input_dir, root / "middle", root / "output")
            output_workbook = root / "middle" / "00_input" / "combined.xlsx"
            output_config = root / "middle" / "00_input" / "pipeline.generated.yaml"

            sources = materialize_inputs(input_dir, config_path, output_workbook, output_config)

            self.assertEqual(
                sources,
                [
                    {
                        "name": "ALL",
                        "label": "ALL尺码匹配表",
                        "sheet": "ALL尺码匹配",
                        "header_row": 1,
                        "columns": "MODEL,YEAR,SIZE",
                        "file": "ALL.csv",
                    },
                    {
                        "name": "TM",
                        "label": "TM尺码匹配表",
                        "sheet": "TM尺码匹配",
                        "header_row": 1,
                        "columns": "MODEL,YEAR,SIZE",
                        "file": "TM.xlsx",
                    },
                ],
            )
            combined = openpyxl.load_workbook(output_workbook, read_only=True, data_only=True)
            self.assertEqual(combined.sheetnames, ["ALL尺码匹配", "TM尺码匹配"])
            self.assertEqual(combined["ALL尺码匹配"]["B2"].value, "ADX")
            self.assertEqual(combined["TM尺码匹配"]["B2"].value, "Model Y")
            combined.close()

            runtime = yaml.safe_load(output_config.read_text(encoding="utf-8"))
            self.assertEqual(runtime["input"]["sheets"], ["ALL尺码匹配", "TM尺码匹配"])
            self.assertEqual(
                configured_stores(output_config),
                [("ALL", "ALL尺码匹配"), ("TM", "TM尺码匹配")],
            )

    def test_header_aliases_are_applied_before_required_field_validation(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            input_dir = root / "input"
            input_dir.mkdir()
            (input_dir / "TM拆.csv").write_text(
                "S8MAKE,MODEL,YEAR,确认尺码\nAcura,RL,1996-1998,3XXXL-0\n",
                encoding="utf-8",
            )
            config_path = root / "pipeline.yaml"
            write_config(config_path, input_dir, root / "middle", root / "output")

            output_workbook = root / "combined.xlsx"
            materialize_inputs(input_dir, config_path, output_workbook, root / "runtime.yaml")

            combined = openpyxl.load_workbook(output_workbook, read_only=True, data_only=True)
            headers = [cell.value for cell in combined["TM拆尺码匹配"][1]]
            combined.close()
            self.assertEqual(headers[0], "MAKE")

    def test_missing_required_fields_fail_at_input_boundary(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            input_dir = root / "input"
            input_dir.mkdir()
            (input_dir / "broken.csv").write_text(
                "MAKE,MODEL,YEAR\nAcura,ADX,2025\n",
                encoding="utf-8",
            )
            config_path = root / "pipeline.yaml"
            write_config(config_path, input_dir, root / "middle", root / "output")

            with self.assertRaisesRegex(ValueError, "最终尺码"):
                materialize_inputs(
                    input_dir,
                    config_path,
                    root / "combined.xlsx",
                    root / "runtime.yaml",
                )

    def test_scan_cases_groups_all_inputs_into_one_case(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            input_dir = root / "input"
            input_dir.mkdir()
            (input_dir / "ALL.csv").touch()
            (input_dir / "TM.csv").touch()
            config_path = root / "pipeline.yaml"
            write_config(config_path, input_dir, root / "middle", root / "output")
            config = yaml.safe_load(config_path.read_text(encoding="utf-8"))

            [case] = scan_cases(config)

            self.assertEqual(case["case_name"], "combined")
            self.assertEqual([path.name for path in case["input_files"]], ["ALL.csv", "TM.csv"])
            self.assertEqual(case["input_file"], root / "middle" / "00_input" / "combined.xlsx")

    def test_publish_config_keeps_generated_match_source_constraints(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            workspace = root / "workspace"
            (workspace / "config").mkdir(parents=True)
            (workspace / "config" / "size-chart-view.yaml").write_text(
                "size_reference:\n  sheet: ALL尺码\n  data_path: data/generated/size-ref.json\n",
                encoding="utf-8",
            )
            pipeline_config = root / "pipeline.generated.yaml"
            pipeline_config.write_text(
                yaml.safe_dump(
                    {
                        "input": {
                            "match_sources": [
                                {
                                    "name": "新分析0831",
                                    "label": "新分析0831尺码匹配表",
                                    "sheet": "新分析0831尺码匹配",
                                    "header_row": 2,
                                    "columns": "MODEL,YEAR,SIZE",
                                    "file": "新分析0831.csv",
                                }
                            ]
                        }
                    },
                    allow_unicode=True,
                    sort_keys=False,
                ),
                encoding="utf-8",
            )
            workbook_path = root / "combined.xlsx"
            workbook = openpyxl.Workbook()
            workbook.active.title = "新分析0831尺码匹配"
            workbook.save(workbook_path)

            exports_reference = configure_excel_sources(
                workspace, pipeline_config, workbook_path
            )

            view = yaml.safe_load(
                (workspace / "config" / "size-chart-view.yaml").read_text(encoding="utf-8")
            )
            self.assertFalse(exports_reference)
            self.assertEqual(
                view["match_sources"],
                [
                    {
                        "name": "新分析0831",
                        "label": "新分析0831尺码匹配表",
                        "sheet": "新分析0831尺码匹配",
                        "header_row": 2,
                        "columns": "MODEL,YEAR,SIZE",
                    }
                ],
            )


if __name__ == "__main__":
    unittest.main()
