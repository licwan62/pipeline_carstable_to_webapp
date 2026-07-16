from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from run_all import find_project_case, load_config, replace_workspace_transactionally, workspace_fingerprint
from tools.generate_all_html import load_profile
from openpyxl.worksheet.table import Table

from tools.build_user_size_templates import sync_rows
from tools.merge_incremental_workbook import DEFAULT_SHEETS, FITMENT_KEY_COLUMNS, merge_workbooks
from tools.merge_incremental_compress_outputs import merge_compress_outputs
from tools.validate_atom_checks import validate_atom_checks
from tools.validate_incremental_atom_scope import (
    DATASETS,
    atom_scope_keys,
    filter_atoms_by_scope,
    validate_incremental_scope,
)


HEADERS = [*FITMENT_KEY_COLUMNS, "确认尺码"]


def fitment_row(model: str, size: str = "XL") -> list[str]:
    return ["Make", model, f"Make|{model}", "Base", "gen1", "2026", "SUV", "SUV", "", "", size]


def make_workbook(path: Path, rows: list[list[str]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in DEFAULT_SHEETS:
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(HEADERS)
        for row in rows:
            sheet.append(row)
    workbook.save(path)


class IncrementalWorkbookTests(unittest.TestCase):
    def test_merge_appends_new_models(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            base = root / "base.xlsx"
            increment = root / "increment.xlsx"
            output = root / "merged.xlsx"
            make_workbook(base, [fitment_row("Old")])
            make_workbook(increment, [fitment_row("New")])

            summary = merge_workbooks(base, increment, output)

            self.assertTrue(all(item["appended"] == 1 for item in summary.values()))
            workbook = load_workbook(output, read_only=True, data_only=True)
            self.assertEqual(workbook[DEFAULT_SHEETS[0]].max_row, 3)
            workbook.close()

    def test_merge_skips_exact_duplicates(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            base = root / "base.xlsx"
            increment = root / "increment.xlsx"
            output = root / "merged.xlsx"
            make_workbook(base, [fitment_row("Same")])
            make_workbook(increment, [fitment_row("Same")])

            summary = merge_workbooks(base, increment, output)

            self.assertTrue(all(item["duplicates"] == 1 for item in summary.values()))

    def test_merge_rejects_changed_existing_fitment(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            base = root / "base.xlsx"
            increment = root / "increment.xlsx"
            output = root / "merged.xlsx"
            make_workbook(base, [fitment_row("Same", "XL")])
            make_workbook(increment, [fitment_row("Same", "2XL")])

            with self.assertRaisesRegex(ValueError, "相同车型键"):
                merge_workbooks(base, increment, output)

    def test_merge_compress_outputs_appends_and_deduplicates(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            base_root = root / "base"
            incremental_root = root / "incremental"
            columns = ["压缩类型", "MAKE", "MODEL", "YEAR", "BACKSIZE"]
            old_row = ["非皮卡", "Make", "Old", "2020", "XL"]
            new_row = ["非皮卡", "Make", "New", "2026", "XL"]
            for dataset in DEFAULT_SHEETS:
                base_dir = base_root / f"0706_{dataset}" / "compress"
                increment_dir = incremental_root / f"0716_{dataset}" / "compress"
                base_dir.mkdir(parents=True)
                increment_dir.mkdir(parents=True)
                suffix = "原子事实表.tsv"
                pd.DataFrame([old_row, old_row], columns=columns).to_csv(
                    base_dir / f"0706_{dataset}_{suffix}", sep="\t", index=False, encoding="utf-8-sig"
                )
                pd.DataFrame([old_row, new_row], columns=columns).to_csv(
                    increment_dir / f"0716_{dataset}_{suffix}", sep="\t", index=False, encoding="utf-8-sig"
                )

            summary = merge_compress_outputs(base_root, incremental_root)

            for dataset in DEFAULT_SHEETS:
                self.assertEqual(summary[dataset]["原子事实表.tsv"], (2, 2, 3))


class AtomCheckTests(unittest.TestCase):
    def write_check(self, path: Path, result: str) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        atom_type = "皮卡" if "皮卡原子检查" in path.name and "非皮卡" not in path.name else "非皮卡"
        prefix = path.name.removesuffix(f"_{atom_type}原子检查.tsv")
        atom_path = path.parent.parent / "compress" / f"{prefix}_原子事实表.tsv"
        atom_path.parent.mkdir(parents=True, exist_ok=True)
        atom_path.write_text("压缩类型\n" f"{atom_type}\n", encoding="utf-8-sig")
        path.write_text(
            "检查结果\tMAKE\tMODEL\tYEAR\n" f"{result}\tMake\tModel\t2026\n",
            encoding="utf-8-sig",
        )

    def test_atom_check_accepts_only_ok(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            self.write_check(root / "sample" / "check" / "sample_非皮卡原子检查.tsv", "OK")
            checked, errors = validate_atom_checks(root)
            self.assertEqual(checked, 1)
            self.assertEqual(errors, [])

    def test_atom_check_rejects_miss(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            self.write_check(root / "sample" / "check" / "sample_皮卡原子检查.tsv", "MISS")
            _, errors = validate_atom_checks(root)
            self.assertTrue(any("MISS" in error for error in errors))

    def test_incremental_scope_includes_only_matching_model_size_keys(self) -> None:
        incremental = pd.DataFrame(
            [{"MAKE": "Make", "MODEL": "Target", "BACKSIZE": "XL", "YEAR": "2026"}]
        )
        full = pd.DataFrame(
            [
                {"MAKE": "Make", "MODEL": "Target", "BACKSIZE": "XL", "YEAR": "2026"},
                {"MAKE": "Make", "MODEL": "Target", "BACKSIZE": "XL", "YEAR": "2020"},
                {"MAKE": "Make", "MODEL": "Target", "BACKSIZE": "2XL", "YEAR": "2020"},
                {"MAKE": "Make", "MODEL": "Other", "BACKSIZE": "XL", "YEAR": "2020"},
            ]
        )

        scoped = filter_atoms_by_scope(full, atom_scope_keys(incremental))

        self.assertEqual(scoped["YEAR"].tolist(), ["2026", "2020"])

    def test_scoped_validator_checks_new_and_overlapping_history(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            incremental_root = root / "incremental"
            full_root = root / "full"
            atom_columns = [
                "压缩类型", "MAKE", "MODEL", "YEAR", "VERSION", "CONST", "CAB", "BED_FT", "BACKSIZE"
            ]
            incremental_rows = [["非皮卡", "Make", "Target", "2026", "", "SUV", "", "", "XL"]]
            full_rows = [
                incremental_rows[0],
                ["非皮卡", "Make", "Target", "2020", "", "SUV", "", "", "XL"],
                ["非皮卡", "Make", "Other", "2020", "", "SUV", "", "", "XL"],
            ]
            compress_columns = ["MAKE", "MODEL", "YEAR", "VERSION", "CONST", "BACKSIZE"]
            compress_rows = [["Make", "Target", "2020/2026", "", "SUV", "XL"]]
            for dataset in DATASETS:
                increment_dir = incremental_root / f"delta_{dataset}" / "compress"
                complete_dir = full_root / f"base_{dataset}" / "compress"
                increment_dir.mkdir(parents=True)
                complete_dir.mkdir(parents=True)
                pd.DataFrame(incremental_rows, columns=atom_columns).to_csv(
                    increment_dir / f"delta_{dataset}_原子事实表.tsv",
                    sep="\t", index=False, encoding="utf-8-sig",
                )
                pd.DataFrame(full_rows, columns=atom_columns).to_csv(
                    complete_dir / f"base_{dataset}_原子事实表.tsv",
                    sep="\t", index=False, encoding="utf-8-sig",
                )
                pd.DataFrame(compress_rows, columns=compress_columns).to_csv(
                    complete_dir / f"base_{dataset}_非皮卡高度压缩表.tsv",
                    sep="\t", index=False, encoding="utf-8-sig",
                )

            compress_repo = Path(__file__).resolve().parents[2] / "compress_to_size_chart"
            scoped, overlap, errors = validate_incremental_scope(
                incremental_root=incremental_root,
                full_root=full_root,
                compress_repo=compress_repo,
            )

            self.assertEqual(scoped, 6)
            self.assertEqual(overlap, 3)
            self.assertEqual(errors, [])


class UserSizeSyncTests(unittest.TestCase):
    def test_sync_preserves_manual_rows_and_adds_formula_rows(self) -> None:
        existing_workbook = Workbook()
        existing_sheet = existing_workbook.active
        existing_sheet.title = "非皮卡压缩表"
        existing_sheet.append(["MAKE", "MODEL", "SIZE", "CALC"])
        existing_sheet.append(["Make", "Old", "MANUAL", "=A2"])
        existing_sheet.add_table(Table(displayName="非皮卡压缩表", ref="A1:D2"))

        template_workbook = Workbook()
        template_sheet = template_workbook.active
        template_sheet.title = "非皮卡压缩表"
        template_sheet.append(["MAKE", "MODEL", "SIZE", "CALC"])
        template_sheet.append(["", "", "=A2&B2", "=B2"])

        stats = sync_rows(
            existing_sheet,
            template_sheet,
            [{"MAKE": "Make", "MODEL": "Old"}, {"MAKE": "Make", "MODEL": "New"}],
            {"MAKE", "MODEL"},
        )

        self.assertEqual(
            stats,
            {"preserved": 1, "added": 1, "removed": 0, "formula_cells_refreshed": 1},
        )
        self.assertEqual(existing_sheet["C2"].value, "MANUAL")
        self.assertEqual(existing_sheet["C3"].value, "=A3&B3")
        self.assertEqual(existing_sheet["D2"].value, "=B2")
        self.assertEqual(existing_sheet["D3"].value, "=B3")

    def test_manual_workbook_directory_is_excluded_from_workspace_fingerprint(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            middle = root / "data" / "middle"
            manual = middle / "02_user_size_workbooks"
            manual.mkdir(parents=True)
            stable = middle / "01_compress.tsv"
            workbook = manual / "ALL_用户尺码模板.xlsx"
            stable.write_text("stable", encoding="utf-8")
            workbook.write_text("before", encoding="utf-8")
            before = workspace_fingerprint([middle], ignored_roots=(manual,))

            workbook.write_text("after", encoding="utf-8")
            after = workspace_fingerprint([middle], ignored_roots=(manual,))

            self.assertEqual(before, after)


class WorkspacePromotionTests(unittest.TestCase):
    def test_promotes_in_place_and_removes_stale_files(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            current_dirs = {name: root / "data" / name for name in ("input", "middle", "output")}
            staged_dirs = {name: root / "stage" / name for name in current_dirs}
            for name in current_dirs:
                current_dirs[name].mkdir(parents=True)
                staged_dirs[name].mkdir(parents=True)
                (current_dirs[name] / "stale.txt").write_text("stale", encoding="utf-8")
                (staged_dirs[name] / "new.txt").write_text(name, encoding="utf-8")

            replace_workspace_transactionally(
                current_dirs=current_dirs,
                staged_dirs=staged_dirs,
                rollback_root=root / "rollback",
            )

            for name, current in current_dirs.items():
                self.assertEqual((current / "new.txt").read_text(encoding="utf-8"), name)
                self.assertFalse((current / "stale.txt").exists())
                self.assertTrue(current.is_dir())
            self.assertFalse((root / "rollback").exists())


class ConfigLoadingTests(unittest.TestCase):
    def test_pipeline_include_is_deep_merged(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            (root / "advanced.yaml").write_text(
                "paths:\n  input_dir: default\n  output_dir: output\nsteps:\n  one:\n    enabled: true\n",
                encoding="utf-8",
            )
            (root / "main.yaml").write_text(
                "include: advanced.yaml\npaths:\n  input_dir: custom\n",
                encoding="utf-8",
            )

            config = load_config(root / "main.yaml")

            self.assertEqual(config["paths"], {"input_dir": "custom", "output_dir": "output"})
            self.assertTrue(config["steps"]["one"]["enabled"])

    def test_html_profile_extends_common_style(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            (root / "style.yaml").write_text(
                'page_background: "#ffffff"\nfont_size: 18\n', encoding="utf-8"
            )
            (root / "profile.yaml").write_text(
                "extends: style.yaml\nfont_size: 20\nsize_column: SIZE\n", encoding="utf-8"
            )

            profile = load_profile(root / "profile.yaml")

            self.assertEqual(profile["page_background"], "#ffffff")
            self.assertEqual(profile["font_size"], 20)
            self.assertEqual(profile["size_column"], "SIZE")


class CaseDirectoryTests(unittest.TestCase):
    def test_finds_flat_workspace_layout(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            (root / "data/input").mkdir(parents=True)
            (root / "data/input/0706.xlsx").touch()
            (root / "data/middle/02_user_size_workbooks").mkdir(parents=True)
            (root / "data/output/site").mkdir(parents=True)

            case = find_project_case(root)

            self.assertEqual(case["case_name"], "0706")
            self.assertEqual(case["source_dirs"]["middle"], root / "data/middle")
            self.assertEqual(case["source_dirs"]["output"], root / "data/output")

    def test_finds_legacy_nested_archive_layout(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            (root / "data/input").mkdir(parents=True)
            (root / "data/input/0706.xlsx").touch()
            (root / "data/middle/0706/02_user_size_workbooks").mkdir(parents=True)
            (root / "data/output/0706/site").mkdir(parents=True)

            case = find_project_case(root)

            self.assertEqual(case["source_dirs"]["middle"], root / "data/middle/0706")
            self.assertEqual(case["source_dirs"]["output"], root / "data/output/0706")


if __name__ == "__main__":
    unittest.main()
