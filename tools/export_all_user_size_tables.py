from __future__ import annotations

import argparse
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


def workbook_stores(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        stores: set[str] = set()
        for sheet_name in ("非皮卡压缩表", "皮卡压缩表"):
            sheet = workbook[sheet_name]
            headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
            store_column = headers.index("店铺") + 1
            stores.update(
                str(row[0]).strip()
                for row in sheet.iter_rows(min_row=2, min_col=store_column, max_col=store_column, values_only=True)
                if row[0] is not None and str(row[0]).strip()
            )
        return sorted(stores)
    finally:
        workbook.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export all stores from one user-size workbook to TSV tables.")
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--output-root", type=Path, required=True)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    script = Path(__file__).with_name("export_user_size_tables.py")
    for store in workbook_stores(args.workbook):
        output_dir = args.output_root / store
        command = [
            sys.executable,
            str(script),
            "--workbook",
            str(args.workbook),
            "--output-dir",
            str(output_dir),
            "--store",
            store,
        ]
        print(f"[{store}] {' '.join(command)}")
        subprocess.run(command, check=True)


if __name__ == "__main__":
    main()
