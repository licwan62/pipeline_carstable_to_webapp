from __future__ import annotations

import argparse
import copy
import csv
import re
from pathlib import Path
from typing import Any, Iterable

import openpyxl
import yaml


EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
SUPPORTED_SUFFIXES = EXCEL_SUFFIXES | {".csv"}
INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")
DEFAULT_SOURCE_RULE = {
    "name": "{stem}",
    "label": "{stem}尺码匹配表",
    "sheet": "{stem}尺码匹配",
    "header_row": 1,
    "columns": "MODEL,版本,YEAR,TYPE,CAB,BED,L-MM,W-MM,H-MM,长度余量,SIZE",
}
DEFAULT_REQUIRED_FIELDS = ("品牌", "前台车型", "年份区间", "最终尺码")
DEFAULT_REQUIRED_ALIASES = {
    "品牌": ("MAKE", "品牌"),
    "前台车型": ("MODEL", "前台车型"),
    "年份区间": ("YEAR", "年份区间"),
    "最终尺码": ("确认尺码", "自动尺码", "最终尺码", "对应尺码"),
}


class IndentedSafeDumper(yaml.SafeDumper):
    def increase_indent(self, flow: bool = False, indentless: bool = False):
        return super().increase_indent(flow, indentless=False)


def deep_merge(base: dict[str, Any], overrides: dict[str, Any]) -> dict[str, Any]:
    result = copy.deepcopy(base)
    for key, value in overrides.items():
        if isinstance(value, dict) and isinstance(result.get(key), dict):
            result[key] = deep_merge(result[key], value)
        else:
            result[key] = copy.deepcopy(value)
    return result


def load_config(config_path: Path, loading: tuple[Path, ...] = ()) -> dict[str, Any]:
    config_path = config_path.resolve()
    if config_path in loading:
        chain = " -> ".join(str(path) for path in (*loading, config_path))
        raise ValueError(f"Circular config include: {chain}")
    raw = yaml.safe_load(config_path.read_text(encoding="utf-8")) or {}
    includes = raw.pop("include", [])
    if isinstance(includes, str):
        includes = [includes]
    merged: dict[str, Any] = {}
    for included in includes:
        included_path = (config_path.parent / str(included)).resolve()
        merged = deep_merge(merged, load_config(included_path, (*loading, config_path)))
    return deep_merge(merged, raw)


def input_patterns(config: dict[str, Any]) -> list[str]:
    file_rules = config.get("file_rules") or {}
    configured = file_rules.get("input_patterns", file_rules.get("input_pattern", ["*.xlsx", "*.csv"]))
    if isinstance(configured, str):
        configured = [configured]
    if not isinstance(configured, list) or not configured:
        raise ValueError("file_rules.input_patterns must be a non-empty string or list")
    patterns = [str(pattern).strip() for pattern in configured if str(pattern).strip()]
    if not patterns:
        raise ValueError("file_rules.input_patterns must contain at least one pattern")
    return patterns


def discover_input_files(input_dir: Path, config: dict[str, Any]) -> list[Path]:
    if not input_dir.is_dir():
        raise FileNotFoundError(f"Input directory does not exist: {input_dir}")
    discovered: dict[Path, Path] = {}
    for pattern in input_patterns(config):
        for path in input_dir.glob(pattern):
            if not path.is_file() or path.name.startswith(("~$", ".")):
                continue
            if path.suffix.casefold() not in SUPPORTED_SUFFIXES:
                continue
            discovered[path.resolve()] = path.resolve()
    return sorted(discovered.values(), key=lambda path: path.name.casefold())


def render_template(value: Any, path: Path) -> str:
    variables = {
        "stem": path.stem,
        "filename": path.name,
        "suffix": path.suffix.lstrip("."),
    }
    try:
        return str(value).format(**variables).strip()
    except KeyError as exc:
        raise ValueError(f"Unknown input.match_source template variable: {exc.args[0]}") from exc


def safe_sheet_name(value: str) -> str:
    cleaned = INVALID_SHEET_CHARS.sub("_", value).strip().strip("'")
    if not cleaned:
        raise ValueError("Generated worksheet name is empty")
    return cleaned[:31]


def build_match_source(path: Path, config: dict[str, Any]) -> dict[str, Any]:
    input_config = config.get("input") or {}
    configured_rule = input_config.get("match_source") or {}
    if not isinstance(configured_rule, dict):
        raise ValueError("input.match_source must be a mapping")
    rule = {**DEFAULT_SOURCE_RULE, **configured_rule}
    source = {
        "name": render_template(rule["name"], path),
        "label": render_template(rule["label"], path),
        "sheet": safe_sheet_name(render_template(rule["sheet"], path)),
        "header_row": int(rule.get("header_row", 1)),
        "columns": str(rule.get("columns", DEFAULT_SOURCE_RULE["columns"])),
    }
    if not source["name"] or not source["label"]:
        raise ValueError(f"Generated source name/label is empty for {path.name}")
    if source["header_row"] < 1:
        raise ValueError("input.match_source.header_row must be at least 1")
    source["file"] = path.name
    return source


def select_excel_sheet(workbook: Any, path: Path, source: dict[str, Any], config: dict[str, Any]) -> str:
    rule = (config.get("input") or {}).get("match_source") or {}
    configured = rule.get("source_sheet")
    if configured:
        requested = render_template(configured, path)
        if requested not in workbook.sheetnames:
            raise KeyError(
                f"Configured worksheet '{requested}' does not exist in {path.name}; "
                f"available: {workbook.sheetnames}"
            )
        return requested

    candidates = [source["sheet"], source["name"], path.stem, source["label"]]
    for candidate in candidates:
        if candidate in workbook.sheetnames:
            return candidate
        if candidate.endswith("表") and candidate[:-1] in workbook.sheetnames:
            return candidate[:-1]
        if f"{candidate}表" in workbook.sheetnames:
            return f"{candidate}表"
    if len(workbook.sheetnames) == 1:
        return workbook.sheetnames[0]
    raise ValueError(
        f"Cannot select a worksheet from {path.name}; generated source sheet is '{source['sheet']}', "
        f"available: {workbook.sheetnames}. Set input.match_source.source_sheet if needed."
    )


def csv_rows(path: Path, config: dict[str, Any]) -> Iterable[list[str]]:
    csv_config = (config.get("input") or {}).get("csv") or {}
    encoding = str(csv_config.get("encoding", "utf-8-sig"))
    configured_delimiter = csv_config.get("delimiter")
    with path.open("r", encoding=encoding, newline="") as handle:
        if configured_delimiter:
            reader = csv.reader(handle, delimiter=str(configured_delimiter))
        else:
            sample = handle.read(65536)
            handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
            except csv.Error:
                dialect = csv.excel
            reader = csv.reader(handle, dialect)
        yield from reader


def header_aliases(config: dict[str, Any]) -> dict[str, str]:
    configured = (config.get("input") or {}).get("header_aliases") or {}
    if not isinstance(configured, dict):
        raise ValueError("input.header_aliases must be a mapping")
    return {
        str(source).strip(): str(target).strip()
        for source, target in configured.items()
        if str(source).strip() and str(target).strip()
    }


def normalize_headers(headers: Iterable[Any], config: dict[str, Any], path: Path) -> list[Any]:
    aliases = header_aliases(config)
    normalized = [aliases.get(str(value).strip(), value) if value is not None else value for value in headers]
    populated = [str(value).strip() for value in normalized if value is not None and str(value).strip()]
    duplicates = sorted({value for value in populated if populated.count(value) > 1})
    if duplicates:
        raise ValueError(f"Header aliases produce duplicate columns in {path.name}: {duplicates}")
    return normalized


def validate_headers(headers: Iterable[Any], config: dict[str, Any], path: Path) -> None:
    available = {str(value).strip() for value in headers if value is not None and str(value).strip()}
    input_config = config.get("input") or {}
    required = input_config.get("required_fields", DEFAULT_REQUIRED_FIELDS)
    if isinstance(required, str):
        required = [required]
    if not isinstance(required, list | tuple):
        raise ValueError("input.required_fields must be a string or list")

    configured_columns = config.get("columns") or {}
    missing: list[str] = []
    for field in required:
        canonical = str(field).strip()
        candidates = configured_columns.get(canonical, DEFAULT_REQUIRED_ALIASES.get(canonical, [canonical]))
        if isinstance(candidates, str):
            candidates = [candidates]
        accepted = {canonical, *(str(candidate).strip() for candidate in candidates)}
        if not available.intersection(accepted):
            missing.append(f"{canonical} ({'/'.join(sorted(accepted))})")
    if missing:
        raise ValueError(f"Input source {path.name} is missing required fields: {', '.join(missing)}")


def copy_rows(rows: Iterable[Iterable[Any]], target: Any, config: dict[str, Any], path: Path) -> int:
    iterator = iter(rows)
    try:
        try:
            headers = normalize_headers(next(iterator), config, path)
        except StopIteration as exc:
            raise ValueError(f"Input source is empty: {path.name}") from exc
        validate_headers(headers, config, path)
        target.append(headers)

        row_count = 0
        for row in iterator:
            target.append(list(row))
            row_count += 1
        if row_count == 0:
            raise ValueError(f"Input source has a header but no data rows: {path.name}")
        return row_count
    finally:
        close = getattr(iterator, "close", None)
        if close is not None:
            close()


def materialize_inputs(
    input_dir: Path,
    config_path: Path,
    output_workbook: Path,
    output_config: Path,
) -> list[dict[str, Any]]:
    config = load_config(config_path)
    files = discover_input_files(input_dir, config)
    if not files:
        patterns = ", ".join(input_patterns(config))
        raise FileNotFoundError(f"No supported input files found in {input_dir} (patterns: {patterns})")

    sources = [build_match_source(path, config) for path in files]
    names = [str(source["name"]) for source in sources]
    sheets = [str(source["sheet"]) for source in sources]
    if len(names) != len(set(names)):
        raise ValueError(f"Input filenames produce duplicate match source names: {names}")
    if len(sheets) != len(set(sheets)):
        raise ValueError(f"Input filenames produce duplicate worksheet names: {sheets}")

    combined = openpyxl.Workbook(write_only=True)
    for path, source in zip(files, sources, strict=True):
        target_sheet = combined.create_sheet(str(source["sheet"]))
        if path.suffix.casefold() in EXCEL_SUFFIXES:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                selected_sheet = select_excel_sheet(workbook, path, source, config)
                row_count = copy_rows(
                    workbook[selected_sheet].iter_rows(values_only=True), target_sheet, config, path
                )
            finally:
                workbook.close()
        else:
            row_count = copy_rows(csv_rows(path, config), target_sheet, config, path)
        print(
            f"[input-source] {path.name} -> {source['name']} / {source['sheet']} "
            f"({row_count} data rows)"
        )

    output_workbook.parent.mkdir(parents=True, exist_ok=True)
    combined.save(output_workbook)

    runtime_config = dict(config)
    runtime_input = dict(runtime_config.get("input") or {})
    runtime_input["sheets"] = sheets
    runtime_input["match_sources"] = sources
    runtime_config["input"] = runtime_input
    output_config.parent.mkdir(parents=True, exist_ok=True)
    with output_config.open("w", encoding="utf-8", newline="\n") as handle:
        yaml.dump(
            runtime_config,
            handle,
            Dumper=IndentedSafeDumper,
            allow_unicode=True,
            sort_keys=False,
        )
    print(f"Materialized {len(sources)} input source(s): {output_workbook}")
    print(f"Runtime config: {output_config}")
    return sources


def main() -> None:
    parser = argparse.ArgumentParser(description="Combine configured Excel/CSV inputs and generate match_sources.")
    parser.add_argument("--input-dir", type=Path, required=True)
    parser.add_argument("--config", type=Path, required=True)
    parser.add_argument("--output-workbook", type=Path, required=True)
    parser.add_argument("--output-config", type=Path, required=True)
    args = parser.parse_args()
    materialize_inputs(
        args.input_dir.resolve(),
        args.config.resolve(),
        args.output_workbook.resolve(),
        args.output_config.resolve(),
    )


if __name__ == "__main__":
    main()
