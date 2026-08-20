from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def table_rows(workbook, table_name: str) -> list[dict[str, str]]:
    for sheet in workbook.worksheets:
        if table_name not in sheet.tables:
            continue
        table = sheet.tables[table_name]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [clean(sheet.cell(min_row, col).value) for col in range(min_col, max_col + 1)]
        return [
            {header: clean(sheet.cell(row, col).value) for header, col in zip(headers, range(min_col, max_col + 1)) if header}
            for row in range(min_row + 1, max_row + 1)
            if any(sheet.cell(row, col).value is not None for col in range(min_col, max_col + 1))
        ]
    raise KeyError(f"Excel table not found: {table_name}")


def extract(template: Path) -> dict:
    workbook = load_workbook(template, data_only=True, read_only=False)
    try:
        size_rows = table_rows(workbook, "ALL尺码表")
        type_rows = table_rows(workbook, "TYPE缩写表")
        sort_sheet = workbook["排序规则"]
        return {
            "version": 1,
            "size": {row["内部尺码"]: {"category": row.get("分类", ""), "generic": row.get("通用尺码", "")} for row in size_rows},
            "category_order": [clean(sort_sheet.cell(row, 3).value) for row in range(2, sort_sheet.max_row + 1) if clean(sort_sheet.cell(row, 3).value) and not clean(sort_sheet.cell(row, 3).value).startswith("=")],
            "pickup_front": table_rows(workbook, "皮卡前台名表"),
            "model_abbreviations": {row["长MODEL"]: row["短MODEL"] for row in table_rows(workbook, "MODEL缩写表") if row.get("长MODEL")},
            "type_abbreviations": [{"car": row.get("CAR", ""), "long": row.get("LONG-TYPE", ""), "short": row.get("SHORT-TYPE", "")} for row in type_rows if row.get("LONG-TYPE")],
            "cab_abbreviations": {row["LONG-CAB"]: row["SHORT-CAB"] for row in table_rows(workbook, "CAB缩写表") if row.get("LONG-CAB")},
            "ai_cache": {"model": {}, "type": {}, "cab": {}},
        }
    finally:
        workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract user-size lookup rules from the legacy Excel template.")
    parser.add_argument("--template", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(extract(args.template), ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"Rules written: {args.output}")


if __name__ == "__main__":
    main()
