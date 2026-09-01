from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict, deque
from copy import copy
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import Table
import yaml


DEFAULT_TEMPLATE = Path("data/template/尺码适配表.xlsx")
NON_PICKUP_SHEET = "非皮卡压缩表"
PICKUP_SHEET = "皮卡压缩表"
DATA_ROW = 2
TEXT_NUMBER_FORMAT = "@"
TEXT_COLUMNS_BY_SHEET = {
    NON_PICKUP_SHEET: {"店铺", "CAR", "MAKE", "MODEL", "YEAR", "VERSION", "CONST", "BACKSIZE"},
    PICKUP_SHEET: {"店铺", "MAKE", "MODEL", "YEAR", "VERSION", "CAB", "BED", "BACKSIZE"},
}


def store_name(input_sheet: str) -> str:
    normalized = input_sheet.strip().upper()
    if normalized.startswith("全") or normalized.startswith("ALL"):
        return "ALL"
    if normalized in {"TM尺码匹配", "TM尺码匹配表"}:
        return "TM"
    if normalized in {"HNT尺码匹配", "HNT尺码匹配表"}:
        return "HNT"
    label = re.sub(r"尺码匹配表?$", "", input_sheet.strip(), flags=re.IGNORECASE)
    return re.sub(r"[^0-9A-Za-z_\-\u4e00-\u9fff]+", "_", label).strip("_") or "STORE"


def configured_stores(config_path: Path) -> list[tuple[str, str]]:
    config = yaml.safe_load(config_path.read_text(encoding="utf-8")) or {}
    input_config = config.get("input", {})
    match_sources = input_config.get("match_sources") or []
    if match_sources:
        stores = []
        for source in match_sources:
            if not isinstance(source, dict) or not source.get("name") or not source.get("sheet"):
                raise ValueError(f"Invalid input.match_sources entry in {config_path}: {source!r}")
            stores.append((str(source["name"]), str(source["sheet"])))
    else:
        sheets = input_config.get("sheets", [])
        stores = [(store_name(str(sheet)), str(sheet)) for sheet in sheets]
    if not stores:
        raise ValueError(f"No input.match_sources or input.sheets configured in {config_path}")
    labels = [store for store, _ in stores]
    if len(labels) != len(set(labels)):
        raise ValueError(f"Configured inputs produce duplicate store names: {labels}")
    return stores


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return list(csv.DictReader(file, delimiter="\t"))


def table_for_sheet(sheet) -> Table:
    if len(sheet.tables) != 1:
        raise ValueError(f"Sheet '{sheet.title}' must contain exactly one Excel table.")
    table_name = next(iter(sheet.tables.keys()))
    return sheet.tables[table_name]


def copy_cell_template(source, target) -> None:
    target._style = copy(source._style)
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
    target.number_format = source.number_format
    target.data_type = source.data_type
    if isinstance(source.value, ArrayFormula):
        target.value = ArrayFormula(target.coordinate, source.value.text)
    else:
        target.value = copy(source.value)


def cell_snapshot(cell) -> dict[str, object]:
    value: object
    if isinstance(cell.value, ArrayFormula):
        value = ("array_formula", cell.value.text)
    else:
        value = copy(cell.value)
    return {
        "value": value,
        "coordinate": cell.coordinate,
        "style": copy(cell._style),
        "hyperlink": copy(cell.hyperlink),
        "comment": copy(cell.comment),
    }


def apply_cell_snapshot(snapshot: dict[str, object], target) -> None:
    target._style = copy(snapshot["style"])
    target.hyperlink = copy(snapshot["hyperlink"])
    target.comment = copy(snapshot["comment"])
    value = snapshot["value"]
    if isinstance(value, tuple) and value[0] == "array_formula":
        target.value = ArrayFormula(target.coordinate, value[1])
    elif isinstance(value, str) and value.startswith("="):
        target.value = Translator(value, origin=str(snapshot["coordinate"])).translate_formula(target.coordinate)
    else:
        target.value = copy(value)


def snapshot_is_formula(snapshot: dict[str, object]) -> bool:
    value = snapshot["value"]
    return (isinstance(value, tuple) and value[0] == "array_formula") or (
        isinstance(value, str) and value.startswith("=")
    )


def text_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.time() == time.min:
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    return str(value)


def write_text_cell(cell, value: Any) -> None:
    cell.value = text_value(value)
    cell.data_type = "s"
    cell.number_format = TEXT_NUMBER_FORMAT


def reset_sheet_to_template_rows(sheet) -> None:
    if sheet.max_row > DATA_ROW:
        sheet.delete_rows(DATA_ROW + 1, sheet.max_row - DATA_ROW)


def expand_table(sheet, row_count: int) -> None:
    table = table_for_sheet(sheet)
    end_row = max(DATA_ROW, DATA_ROW + row_count - 1)
    end_column = table.ref.split(":")[-1].rstrip("0123456789")
    table.ref = f"A1:{end_column}{end_row}"


def sheet_headers(sheet) -> list[str]:
    return [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]


def write_rows(sheet, rows: list[dict[str, str]], value_columns: set[str]) -> None:
    headers = sheet_headers(sheet)
    template_cells = [sheet.cell(DATA_ROW, column) for column in range(1, len(headers) + 1)]
    reset_sheet_to_template_rows(sheet)

    for row_index, row in enumerate(rows, start=DATA_ROW):
        if row_index > DATA_ROW:
            sheet.insert_rows(row_index)

        for column_index, header in enumerate(headers, start=1):
            target = sheet.cell(row_index, column_index)
            copy_cell_template(template_cells[column_index - 1], target)
            if header in value_columns:
                write_text_cell(target, row.get(header, ""))

    expand_table(sheet, len(rows))


def normalize_key_value(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\u00a0", " ").replace("\u200b", "").strip()


def sync_rows(sheet, template_sheet, rows: list[dict[str, str]], value_columns: set[str]) -> dict[str, int]:
    headers = sheet_headers(sheet)
    if headers != sheet_headers(template_sheet):
        raise ValueError(f"Sheet '{sheet.title}' headers do not match the current template.")

    key_headers = [header for header in headers if header in value_columns]
    existing: dict[tuple[str, ...], deque[list[dict[str, object]]]] = defaultdict(deque)
    for row_index in range(DATA_ROW, sheet.max_row + 1):
        key = tuple(
            normalize_key_value(sheet.cell(row_index, headers.index(header) + 1).value)
            for header in key_headers
        )
        existing[key].append(
            [cell_snapshot(sheet.cell(row_index, column)) for column in range(1, len(headers) + 1)]
        )

    template_snapshots = [
        cell_snapshot(template_sheet.cell(DATA_ROW, column)) for column in range(1, len(headers) + 1)
    ]
    if sheet.max_row >= DATA_ROW:
        sheet.delete_rows(DATA_ROW, sheet.max_row - DATA_ROW + 1)

    preserved = added = formula_cells_refreshed = 0
    for row_index, row in enumerate(rows, start=DATA_ROW):
        key = tuple(normalize_key_value(row.get(header, "")) for header in key_headers)
        matched_existing = bool(existing[key])
        snapshots = existing[key].popleft() if matched_existing else template_snapshots
        preserved += int(matched_existing)
        added += int(not matched_existing)

        for column_index, (header, old_snapshot, template_snapshot) in enumerate(
            zip(headers, snapshots, template_snapshots), start=1
        ):
            snapshot = old_snapshot
            if matched_existing and snapshot_is_formula(template_snapshot):
                manual_size = (
                    header == "SIZE"
                    and not snapshot_is_formula(old_snapshot)
                    and bool(normalize_key_value(old_snapshot["value"]))
                )
                if not manual_size:
                    snapshot = template_snapshot
                    formula_cells_refreshed += 1
            target = sheet.cell(row_index, column_index)
            apply_cell_snapshot(snapshot, target)
            if header in value_columns:
                write_text_cell(target, row.get(header, ""))

    removed = sum(len(items) for items in existing.values())
    expand_table(sheet, len(rows))
    return {
        "preserved": preserved,
        "added": added,
        "removed": removed,
        "formula_cells_refreshed": formula_cells_refreshed,
    }


def non_pickup_rows(rows: list[dict[str, str]], store: str) -> list[dict[str, str]]:
    return [
        {
            "店铺": store,
            "CAR": row.get("CAR", ""),
            "MAKE": row.get("MAKE", ""),
            "MODEL": row.get("MODEL", ""),
            "YEAR": row.get("YEAR", ""),
            "VERSION": row.get("VERSION", ""),
            "CONST": row.get("CONST", ""),
            "BACKSIZE": row.get("BACKSIZE", ""),
        }
        for row in rows
    ]


def pickup_rows(rows: list[dict[str, str]], store: str) -> list[dict[str, str]]:
    return [
        {
            "店铺": store,
            "MAKE": row.get("MAKE", ""),
            "MODEL": row.get("MODEL", ""),
            "YEAR": row.get("YEAR", ""),
            "VERSION": row.get("VERSION", ""),
            "CAB": row.get("CAB", ""),
            "BED": row.get("BED", ""),
            "BACKSIZE": row.get("BACKSIZE", ""),
        }
        for row in rows
    ]


def build_combined_workbook(
    *,
    case_name: str,
    compress_root: Path,
    output_path: Path,
    template_path: Path,
    non_pickup_table_name: str,
    pickup_table_name: str,
    overwrite: bool,
    sync_existing: bool,
    stores: list[tuple[str, str]],
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    all_non_pickup_rows: list[dict[str, str]] = []
    all_pickup_rows: list[dict[str, str]] = []
    for store, input_sheet in stores:
        stem = f"{case_name}_{input_sheet}"
        compress_dir = compress_root / stem / "compress"
        non_pickup_path = compress_dir / f"{stem}_{non_pickup_table_name}.tsv"
        pickup_path = compress_dir / f"{stem}_{pickup_table_name}.tsv"
        all_non_pickup_rows.extend(non_pickup_rows(read_tsv(non_pickup_path), store))
        all_pickup_rows.extend(pickup_rows(read_tsv(pickup_path), store))

    if output_path.exists() and not overwrite:
        if is_compatible_existing_workbook(output_path):
            if sync_existing:
                workbook = load_workbook(output_path, data_only=False)
                template = load_workbook(template_path, data_only=False)
                try:
                    non_pickup_stats = sync_rows(
                        workbook[NON_PICKUP_SHEET], template[NON_PICKUP_SHEET],
                        all_non_pickup_rows, TEXT_COLUMNS_BY_SHEET[NON_PICKUP_SHEET],
                    )
                    pickup_stats = sync_rows(
                        workbook[PICKUP_SHEET], template[PICKUP_SHEET],
                        all_pickup_rows, TEXT_COLUMNS_BY_SHEET[PICKUP_SHEET],
                    )
                    workbook.save(output_path)
                finally:
                    workbook.close()
                    template.close()
                print(f"Synced existing workbook: {output_path} | non-pickup={non_pickup_stats} | pickup={pickup_stats}")
                return output_path
            print(f"Template already exists, keep manual edits: {output_path}")
            return output_path
        print(f"Existing workbook is not based on current template, regenerate: {output_path}")

    workbook = load_workbook(template_path, data_only=False)
    write_rows(
        workbook[NON_PICKUP_SHEET],
        all_non_pickup_rows,
        TEXT_COLUMNS_BY_SHEET[NON_PICKUP_SHEET],
    )
    write_rows(
        workbook[PICKUP_SHEET],
        all_pickup_rows,
        TEXT_COLUMNS_BY_SHEET[PICKUP_SHEET],
    )
    workbook.save(output_path)
    return output_path


def is_compatible_existing_workbook(path: Path) -> bool:
    try:
        workbook = load_workbook(path, read_only=False, data_only=False)
    except Exception:
        return False
    try:
        required_sheets = {
            "ref-ALL尺码表",
            "排序规则",
            "皮卡前台名",
            "MODEL缩写",
            "TYPE缩写",
            "CAB缩写",
            NON_PICKUP_SHEET,
            PICKUP_SHEET,
        }
        if not required_sheets.issubset(set(workbook.sheetnames)):
            return False
        for sheet_name in [NON_PICKUP_SHEET, PICKUP_SHEET]:
            sheet = workbook[sheet_name]
            if sheet_name not in sheet.tables:
                return False
        return True
    finally:
        workbook.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build one combined user-size workbook from the Excel template.")
    parser.add_argument("--case-name", required=True)
    parser.add_argument("--compress-root", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument("--config", type=Path, required=True, help="Pipeline YAML containing input.sheets.")
    parser.add_argument("--non-pickup-table-name", default="非皮卡高度压缩表")
    parser.add_argument("--pickup-table-name", default="皮卡高度压缩表")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing middle workbooks.")
    parser.add_argument(
        "--sync-existing",
        action="store_true",
        help="Reconcile compressed rows while preserving unchanged rows and manual SIZE edits.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_path = build_combined_workbook(
        case_name=args.case_name,
        compress_root=args.compress_root,
        output_path=args.output,
        template_path=args.template,
        non_pickup_table_name=args.non_pickup_table_name,
        pickup_table_name=args.pickup_table_name,
        overwrite=args.overwrite,
        sync_existing=args.sync_existing,
        stores=configured_stores(args.config),
    )
    print(f"Template ready: {output_path}")
    print("请用 Excel/WPS 打开工作簿，确认公式计算出的用户尺码 SIZE 后保存；需要时可人工调整模板表。")


if __name__ == "__main__":
    main()
