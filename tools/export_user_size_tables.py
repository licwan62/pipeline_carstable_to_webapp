from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


NULL_SIZE = "无可用尺码"
STORES = ["ALL", "TM", "HNT"]
NON_PICKUP_SHEET = "非皮卡压缩表"
PICKUP_SHEET = "皮卡压缩表"


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def table_records(workbook, sheet_name: str, table_name: str) -> list[dict[str, str]]:
    sheet = workbook[sheet_name]
    table = sheet.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [clean(sheet.cell(min_row, column).value) for column in range(min_col, max_col + 1)]
    records: list[dict[str, str]] = []
    for row_index in range(min_row + 1, max_row + 1):
        record = {
            header: clean(sheet.cell(row_index, column).value)
            for header, column in zip(headers, range(min_col, max_col + 1))
            if header
        }
        if any(record.values()):
            records.append(record)
    return records


def sheet_records(workbook, sheet_name: str) -> list[dict[str, str]]:
    sheet = workbook[sheet_name]
    headers = [clean(cell.value) for cell in sheet[1]]
    records: list[dict[str, str]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = {header: clean(value) for header, value in zip(headers, row) if header}
        if any(record.values()):
            records.append(record)
    return records


def first_non_empty(*values: str) -> str:
    for value in values:
        if clean(value):
            return clean(value)
    return ""


def build_size_maps(workbook) -> tuple[dict[str, str], dict[str, str]]:
    rows = table_records(workbook, "ref-ALL尺码表", "ALL尺码表")
    size_map = {row.get("内部尺码", ""): row.get("通用尺码", "") for row in rows if row.get("内部尺码")}
    category_map = {row.get("内部尺码", ""): row.get("分类", "") for row in rows if row.get("内部尺码")}
    return size_map, category_map


def build_simple_map(workbook, sheet_name: str, table_name: str, key: str, value: str) -> dict[str, str]:
    rows = table_records(workbook, sheet_name, table_name)
    return {row.get(key, ""): row.get(value, "") for row in rows if row.get(key)}


def build_type_rows(workbook) -> list[dict[str, str]]:
    return table_records(workbook, "TYPE缩写", "TYPE缩写表")


def build_pickup_title_rows(workbook) -> list[dict[str, str]]:
    return table_records(workbook, "皮卡前台名", "皮卡前台名表")


def display_size(row: dict[str, str], size_map: dict[str, str]) -> str:
    explicit = clean(row.get("SIZE", ""))
    if explicit:
        return explicit
    backsize = clean(row.get("BACKSIZE", ""))
    return clean(size_map.get(backsize, ""))


def long_type_for_row(row: dict[str, str], const_counts: dict[str, int]) -> str:
    car = clean(row.get("CAR", ""))
    const = clean(row.get("CONST", ""))
    version = clean(row.get("VERSION", ""))
    if const_counts.get(car, 0) <= 1:
        return version
    return " ".join(part for part in [const, version] if part).strip()


def short_type(long_type: str, car: str, type_rows: list[dict[str, str]]) -> str:
    long_type = clean(long_type)
    car = clean(car)
    if not long_type:
        return ""

    best_score = -1
    best_value = ""
    for row in type_rows:
        row_long = clean(row.get("LONG-TYPE", ""))
        if row_long != long_type:
            continue
        row_car = clean(row.get("CAR", ""))
        score = 2 if row_car == car else 1 if row_car == "" else 0
        if score > best_score:
            best_score = score
            best_value = clean(row.get("SHORT-TYPE", ""))
    return best_value or long_type


def pickup_title(row: dict[str, str], title_rows: list[dict[str, str]]) -> tuple[str, str]:
    make = clean(row.get("MAKE", ""))
    model = clean(row.get("MODEL", ""))
    for title_row in title_rows:
        if clean(title_row.get("MAKE", "")) == make and clean(title_row.get("MODEL", "")) == model:
            return clean(title_row.get("TITLE", "")), clean(title_row.get("DESCRIPTION", ""))
    for title_row in title_rows:
        if clean(title_row.get("MAKE", "")) == make:
            return clean(title_row.get("TITLE", "")), clean(title_row.get("DESCRIPTION", ""))
    return f"{make} {model}".strip(), ""


def export_non_pickup(workbook, output_path: Path) -> int:
    size_map, category_map = build_size_maps(workbook)
    model_map = build_simple_map(workbook, "MODEL缩写", "MODEL缩写表", "长MODEL", "短MODEL")
    type_rows = build_type_rows(workbook)
    rows = sheet_records(workbook, NON_PICKUP_SHEET)

    const_sets: dict[str, set[str]] = {}
    for row in rows:
        car = clean(row.get("CAR", ""))
        const = clean(row.get("CONST", ""))
        const_sets.setdefault(car, set())
        if const:
            const_sets[car].add(const)
    const_counts = {car: len(values) for car, values in const_sets.items()}

    output_rows: list[dict[str, str]] = []
    for row in rows:
        size = display_size(row, size_map)
        if not size or size == NULL_SIZE:
            continue
        long_type = first_non_empty(row.get("LONG-TYPE", ""), long_type_for_row(row, const_counts))
        output_rows.append(
            {
                "店铺": row.get("店铺", ""),
                "CAR": row.get("CAR", ""),
                "MAKE": row.get("MAKE", ""),
                "MODEL": row.get("MODEL", ""),
                "YEAR": row.get("YEAR", ""),
                "VERSION": row.get("VERSION", ""),
                "CONST": row.get("CONST", ""),
                "BACKSIZE": row.get("BACKSIZE", ""),
                "CATAGORY": first_non_empty(row.get("CATAGORY", ""), category_map.get(row.get("BACKSIZE", ""), "")),
                "LONG-TYPE": long_type,
                "TYPE": first_non_empty(row.get("TYPE", ""), short_type(long_type, row.get("CAR", ""), type_rows)),
                "SHORT-MODEL": first_non_empty(row.get("SHORT-MODEL", ""), model_map.get(row.get("MODEL", ""), row.get("MODEL", ""))),
                "SIZE": size,
            }
        )

    write_tsv(output_path, output_rows)
    return len(output_rows)


def export_pickup(workbook, output_path: Path) -> int:
    size_map, _category_map = build_size_maps(workbook)
    cab_map = build_simple_map(workbook, "CAB缩写", "CAB缩写表", "LONG-CAB", "SHORT-CAB")
    title_rows = build_pickup_title_rows(workbook)
    rows = sheet_records(workbook, PICKUP_SHEET)

    output_rows: list[dict[str, str]] = []
    for row in rows:
        size = display_size(row, size_map)
        if not size or size == NULL_SIZE:
            continue
        title, description = pickup_title(row, title_rows)
        output_rows.append(
            {
                "店铺": row.get("店铺", ""),
                "MAKE": row.get("MAKE", ""),
                "MODEL": row.get("MODEL", ""),
                "YEAR": row.get("YEAR", ""),
                "VERSION": row.get("VERSION", ""),
                "CAB": row.get("CAB", ""),
                "BED": row.get("BED", ""),
                "BACKSIZE": row.get("BACKSIZE", ""),
                "SHORT-CAB": first_non_empty(row.get("SHORT-CAB", ""), cab_map.get(row.get("CAB", ""), row.get("CAB", ""))),
                "TITLE": first_non_empty(row.get("TITLE", ""), title),
                "DESCRIPTION": first_non_empty(row.get("DESCRIPTION", ""), description),
                "SIZE": size,
            }
        )

    write_tsv(output_path, output_rows)
    return len(output_rows)


def write_tsv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(rows[0].keys()) if rows else []
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, delimiter="\t", fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export calculated user-size workbook tables to TSV for HTML generation.")
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook = load_workbook(args.workbook, read_only=False, data_only=True)
    non_pickup_count = export_non_pickup(workbook, args.output_dir / "non_pickup.tsv")
    pickup_count = export_pickup(workbook, args.output_dir / "pickup.tsv")
    print(f"Exported {non_pickup_count} non-pickup row(s), {pickup_count} pickup row(s): {args.output_dir}")
    if non_pickup_count == 0 and pickup_count == 0:
        raise SystemExit("No usable SIZE rows after filtering blank and 无可用尺码.")


if __name__ == "__main__":
    main()
