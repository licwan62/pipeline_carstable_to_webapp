from __future__ import annotations

import argparse
import shutil
import sys
import warnings
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula


DEFAULT_SHEETS = ("ALL尺码匹配", "TM尺码匹配", "HNT尺码匹配")
FITMENT_KEY_COLUMNS = (
    "MAKE",
    "MODEL",
    "SUB-MODEL",
    "版本",
    "代际",
    "YEAR",
    "分类",
    "结构",
    "CAB",
    "BED",
)


def normalize_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\u00a0", " ").replace("\u200b", "").strip()


def used_headers(sheet) -> list[str]:
    values = [normalize_value(cell.value) for cell in sheet[1]]
    while values and not values[-1]:
        values.pop()
    if not values or any(not value for value in values):
        raise ValueError(f"工作表 {sheet.title} 的有效表头中存在空列")
    if len(values) != len(set(values)):
        raise ValueError(f"工作表 {sheet.title} 的表头存在重复列")
    return values


def copy_cell(source, target) -> None:
    target._style = copy(source._style)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)
    target.number_format = source.number_format
    if isinstance(source.value, ArrayFormula):
        target.value = ArrayFormula(target.coordinate, source.value.text)
    else:
        target.value = copy(source.value)


def row_is_blank(values: tuple[Any, ...]) -> bool:
    return all(not normalize_value(value) for value in values)


def row_signature(values: tuple[Any, ...]) -> tuple[str, ...]:
    return tuple(normalize_value(value) for value in values)


def key_signature(values: tuple[Any, ...], headers: list[str]) -> tuple[str, ...]:
    indexes = [headers.index(column) for column in FITMENT_KEY_COLUMNS]
    return tuple(normalize_value(values[index]) for index in indexes)


def update_table_range(sheet, header_count: int) -> None:
    if not sheet.tables:
        return
    end_row = max(1, sheet.max_row)
    end_column = get_column_letter(header_count)
    for table in sheet.tables.values():
        table.ref = f"A1:{end_column}{end_row}"


def materialize_formula_values(formula_sheet, values_sheet, width: int) -> None:
    """Keep calculated values stable after openpyxl rewrites the merged workbook."""
    for row_number in range(2, formula_sheet.max_row + 1):
        for column in range(1, width + 1):
            cell = formula_sheet.cell(row_number, column)
            if cell.data_type == "f" or isinstance(cell.value, ArrayFormula):
                cell.value = values_sheet.cell(row_number, column).value


def merge_sheet(base_sheet, base_values_sheet, increment_formula_sheet, increment_values_sheet) -> dict[str, int]:
    base_headers = used_headers(base_sheet)
    increment_headers = used_headers(increment_formula_sheet)
    if increment_headers != base_headers:
        raise ValueError(
            f"工作表 {base_sheet.title} 表头不一致。原项目={base_headers}，增量={increment_headers}"
        )
    missing_keys = [column for column in FITMENT_KEY_COLUMNS if column not in base_headers]
    if missing_keys:
        raise ValueError(f"工作表 {base_sheet.title} 缺少车型键列: {', '.join(missing_keys)}")

    width = len(base_headers)
    materialize_formula_values(base_sheet, base_values_sheet, width)
    existing_signatures: set[tuple[str, ...]] = set()
    existing_keys: set[tuple[str, ...]] = set()
    for row in base_values_sheet.iter_rows(min_row=2, max_col=width, values_only=True):
        values = tuple(row)
        if row_is_blank(values):
            continue
        existing_signatures.add(row_signature(values))
        existing_keys.add(key_signature(values, base_headers))

    appended = 0
    duplicates = 0
    conflicts: list[str] = []
    for row_number in range(2, increment_formula_sheet.max_row + 1):
        data_values = tuple(
            increment_values_sheet.cell(row_number, column).value for column in range(1, width + 1)
        )
        if row_is_blank(data_values):
            continue
        signature = row_signature(data_values)
        key = key_signature(data_values, base_headers)
        if signature in existing_signatures:
            duplicates += 1
            continue
        if key in existing_keys:
            key_text = " | ".join(key)
            conflicts.append(f"{base_sheet.title}!{row_number}: {key_text}")
            continue

        target_row = base_sheet.max_row + 1
        for column in range(1, width + 1):
            source = increment_formula_sheet.cell(row_number, column)
            target = base_sheet.cell(target_row, column)
            copy_cell(source, target)
            if isinstance(source.value, (str, ArrayFormula)) and (
                isinstance(source.value, ArrayFormula) or source.data_type == "f"
            ):
                cached_value = increment_values_sheet.cell(row_number, column).value
                if cached_value is not None:
                    target.value = cached_value
        existing_signatures.add(signature)
        existing_keys.add(key)
        appended += 1

    if conflicts:
        details = "\n".join(f"- {item}" for item in conflicts[:20])
        raise ValueError(
            "增量数据与原项目存在相同车型键但内容不同，已停止合并。"
            "如需修改旧车型，请先走全量更新。\n" + details
        )

    update_table_range(base_sheet, width)
    return {"appended": appended, "duplicates": duplicates}


def merge_workbooks(
    base_path: Path,
    incremental_path: Path,
    output_path: Path,
    sheet_names: tuple[str, ...] = DEFAULT_SHEETS,
) -> dict[str, dict[str, int]]:
    base_path = base_path.resolve()
    incremental_path = incremental_path.resolve()
    output_path = output_path.resolve()
    if base_path == incremental_path:
        raise ValueError("增量文件不能与原项目文件相同")
    if not base_path.is_file():
        raise FileNotFoundError(f"原项目文件不存在: {base_path}")
    if not incremental_path.is_file():
        raise FileNotFoundError(f"增量文件不存在: {incremental_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(base_path, output_path)
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message=".*extension is not supported.*", category=UserWarning)
        base_workbook = load_workbook(output_path, data_only=False)
        base_values_workbook = load_workbook(output_path, data_only=True)
        increment_formula_workbook = load_workbook(incremental_path, data_only=False)
        increment_values_workbook = load_workbook(incremental_path, data_only=True)

    try:
        missing_base = [name for name in sheet_names if name not in base_workbook.sheetnames]
        missing_increment = [name for name in sheet_names if name not in increment_formula_workbook.sheetnames]
        if missing_base:
            raise ValueError(f"原项目缺少工作表: {', '.join(missing_base)}")
        if missing_increment:
            raise ValueError(f"增量文件缺少工作表: {', '.join(missing_increment)}")

        summary: dict[str, dict[str, int]] = {}
        for sheet_name in sheet_names:
            summary[sheet_name] = merge_sheet(
                base_workbook[sheet_name],
                base_values_workbook[sheet_name],
                increment_formula_workbook[sheet_name],
                increment_values_workbook[sheet_name],
            )
        base_workbook.save(output_path)
        return summary
    finally:
        base_workbook.close()
        base_values_workbook.close()
        increment_formula_workbook.close()
        increment_values_workbook.close()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="把新车型增量合并进原项目 Excel。")
    parser.add_argument("--base", type=Path, required=True)
    parser.add_argument("--incremental", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args(argv)

    try:
        summary = merge_workbooks(args.base, args.incremental, args.output)
    except (FileNotFoundError, ValueError, OSError) as exc:
        print(f"增量合并失败: {exc}", file=sys.stderr)
        return 1

    for sheet_name, counts in summary.items():
        print(f"{sheet_name}: 新增 {counts['appended']} 行，跳过重复 {counts['duplicates']} 行")
    print(f"合并完成: {args.output.resolve()}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
