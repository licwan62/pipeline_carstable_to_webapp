from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


REQUIRED_SHEETS = ["非皮卡压缩表", "皮卡压缩表"]
REQUIRED_STORES = {"ALL", "TM", "HNT"}
TEXT_NUMBER_FORMAT = "@"
TEXT_COLUMNS_BY_SHEET = {
    "非皮卡压缩表": {"店铺", "CAR", "MAKE", "MODEL", "YEAR", "VERSION", "CONST", "BACKSIZE"},
    "皮卡压缩表": {"店铺", "MAKE", "MODEL", "YEAR", "VERSION", "CAB", "BED", "BACKSIZE"},
}
MAX_CELL_ERROR_DETAILS = 20


def validate_workbook(path: Path) -> list[str]:
    errors: list[str] = []
    if not path.exists():
        return [f"模板不存在: {path}"]

    workbook = load_workbook(path, read_only=True, data_only=False)
    for sheet_name in REQUIRED_SHEETS:
        if sheet_name not in workbook.sheetnames:
            errors.append(f"{path}: 缺少工作表 {sheet_name}")
            continue

        sheet = workbook[sheet_name]
        header = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        if "SIZE" not in header:
            errors.append(f"{path} / {sheet_name}: 缺少 SIZE 列")
            continue
        if "店铺" not in header:
            errors.append(f"{path} / {sheet_name}: 缺少 店铺 列")
            continue

        store_column = header.index("店铺") + 1
        stores = {
            str(values[0]).strip()
            for values in sheet.iter_rows(
                min_row=2,
                min_col=store_column,
                max_col=store_column,
                values_only=True,
            )
            if values[0] is not None
        }
        missing_stores = sorted(REQUIRED_STORES - stores)
        if missing_stores:
            errors.append(f"{path} / {sheet_name}: 缺少店铺数据 {', '.join(missing_stores)}")

        text_columns = TEXT_COLUMNS_BY_SHEET[sheet_name]
        text_column_indexes = {
            column_index: column_name
            for column_index, column_name in enumerate(header, start=1)
            if column_name in text_columns
        }
        missing_text_columns = sorted(text_columns - set(text_column_indexes.values()))
        if missing_text_columns:
            errors.append(f"{path} / {sheet_name}: 缺少文本列 {', '.join(missing_text_columns)}")
            continue

        cell_errors: list[str] = []
        cell_error_count = 0
        for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            text_cells = [
                (column_index, column_name, row[column_index - 1])
                for column_index, column_name in text_column_indexes.items()
            ]
            if all(cell.value is None for _, _, cell in text_cells):
                continue
            for column_index, column_name, cell in text_cells:
                reasons = []
                if cell.value is not None and not isinstance(cell.value, str):
                    reasons.append(f"值类型={type(cell.value).__name__}")
                if cell.number_format != TEXT_NUMBER_FORMAT:
                    reasons.append(f"数字格式={cell.number_format!r}")
                if reasons:
                    cell_error_count += 1
                    if len(cell_errors) < MAX_CELL_ERROR_DETAILS:
                        coordinate = f"{get_column_letter(column_index)}{row_index}"
                        cell_errors.append(f"{coordinate}({column_name}: {', '.join(reasons)})")
        if cell_error_count:
            detail = "; ".join(cell_errors)
            omitted = cell_error_count - len(cell_errors)
            if omitted:
                detail += f"; 另有 {omitted} 个"
            errors.append(f"{path} / {sheet_name}: {cell_error_count} 个源数据单元格不是文本格式：{detail}")

    return errors


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validate user-size middle workbooks before HTML generation.")
    parser.add_argument("workbooks", nargs="+", type=Path)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    errors: list[str] = []
    for workbook in args.workbooks:
        errors.extend(validate_workbook(workbook))

    if errors:
        print("用户尺码模板未填写完整，暂不生成 HTML：")
        for error in errors:
            print(f"- {error}")
        raise SystemExit(1)

    print("用户尺码模板校验通过。")


if __name__ == "__main__":
    main()
