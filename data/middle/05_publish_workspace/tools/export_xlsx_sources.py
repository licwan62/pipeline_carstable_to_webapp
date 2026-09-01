from __future__ import annotations

import argparse
import hashlib
import json
import re
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "config" / "size-chart-view.yaml"
DIMENSION_ALIASES = {
    "L-MM": (("L-MM", "长_mm"), ("L-CM", "长_cm"), ("L-IN", "长_in", "长")),
    "W-MM": (("W-MM", "宽_mm"), ("W-CM", "宽_cm"), ("W-IN", "宽_in", "宽")),
    "H-MM": (("H-MM", "高_mm"), ("H-CM", "高_cm"), ("H-IN", "高_in", "高")),
}
LEGACY_DIMENSION_FIELDS = {
    alias
    for millimetre_aliases, centimetre_aliases, inch_aliases in DIMENSION_ALIASES.values()
    for alias in (*centimetre_aliases, *inch_aliases)
}
UNAVAILABLE_SIZES = {"无可用尺码", "数据不全"}


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, float):
        return f"{value:.3f}".rstrip("0").rstrip(".")
    return str(value).strip()


def rounded_integer(value: Any, factor: str = "1") -> str:
    """Return a numeric measurement as a conventional half-up integer."""
    text = clean(value)
    if not text:
        return ""
    try:
        number = Decimal(text.replace(",", "")) * Decimal(factor)
    except InvalidOperation:
        return text
    return str(int(number.quantize(Decimal("1"), rounding=ROUND_HALF_UP)))


def first_value(row: dict[str, Any], candidates: tuple[str, ...]) -> Any:
    for candidate in candidates:
        value = row.get(candidate)
        if clean(value):
            return value
    return None


def dimension_mm(row: dict[str, Any], output_field: str) -> str:
    millimetres, centimetres, inches = DIMENSION_ALIASES[output_field]
    for candidates, factor in ((millimetres, "1"), (centimetres, "10"), (inches, "25.4")):
        value = first_value(row, candidates)
        if value is not None:
            return rounded_integer(value, factor)
    return ""


def match_measurement_factor(row: dict[str, Any]) -> str:
    if any(clean(row.get(alias)) for aliases, _, _ in DIMENSION_ALIASES.values() for alias in aliases):
        return "1"
    if any(clean(row.get(alias)) for _, _, aliases in DIMENSION_ALIASES.values() for alias in aliases):
        return "25.4"
    if any(clean(row.get(alias)) for _, aliases, _ in DIMENSION_ALIASES.values() for alias in aliases):
        return "10"
    return "1"


def parse_yaml_config(path: Path) -> dict[str, Any]:
    config: dict[str, Any] = {}
    stack: list[tuple[int, Any]] = [(-1, config)]

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        if not raw_line.strip() or raw_line.lstrip().startswith("#"):
            continue

        indent = len(raw_line) - len(raw_line.lstrip(" "))
        line = raw_line.strip()
        while stack and indent <= stack[-1][0]:
            stack.pop()

        parent = stack[-1][1]
        if line.startswith("- "):
            item_text = line[2:].strip()
            item: dict[str, Any] = {}
            parent.append(item)
            stack.append((indent, item))
            if item_text:
                key, value = split_key_value(item_text)
                item[key] = parse_scalar(value)
            continue

        key, value = split_key_value(line)
        if value == "":
            next_container: Any = [] if key.endswith("sources") else {}
            parent[key] = next_container
            stack.append((indent, next_container))
        else:
            parent[key] = parse_scalar(value)

    return config


def split_key_value(text: str) -> tuple[str, str]:
    if ":" not in text:
        return text, ""
    key, value = text.split(":", 1)
    return key.strip(), value.strip()


def parse_scalar(value: str) -> Any:
    value = value.strip().strip('"').strip("'")
    if value.lower() == "true":
        return True
    if value.lower() == "false":
        return False
    return value


def worksheet_rows(workbook: Any, sheet_name: str, header_row: int) -> tuple[list[str], list[dict[str, str]]]:
    worksheet = workbook[sheet_name]
    raw_headers = [clean(cell) for cell in next(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    headers: list[str] = []
    column_headers: list[str] = []
    seen: dict[str, int] = {}
    for index, header in enumerate(raw_headers, start=1):
        if not header:
            column_headers.append("")
            continue
        if header in seen:
            seen[header] += 1
            header = f"{header}_{seen[header]}"
        else:
            seen[header] = 1
        headers.append(header)
        column_headers.append(header)

    rows: list[dict[str, str]] = []
    for values in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        row: dict[str, str] = {}
        has_value = False
        for header_index, value in enumerate(values):
            if header_index >= len(column_headers):
                break
            header = column_headers[header_index]
            if header:
                text = clean(value)
                row[header] = text
                has_value = has_value or bool(text)
        if has_value:
            rows.append(row)
    return headers, rows


def normalize_size_reference(headers: list[str], rows: list[dict[str, str]]) -> dict[str, Any]:
    normalized_rows = []
    for row in rows:
        if not clean(row.get("内部尺码")):
            continue
        item = {key: value for key, value in row.items() if key not in LEGACY_DIMENSION_FIELDS}
        item["型号"] = clean(row.get("内部尺码"))
        for field in DIMENSION_ALIASES:
            value = dimension_mm(row, field)
            if value:
                item[field.replace("-MM", "_mm").replace("L_", "长_").replace("W_", "宽_").replace("H_", "高_")] = value
        normalized_rows.append(item)

    preferred = ["型号", "分类", "CAB", "长_mm", "宽_mm", "高_mm", "通用尺码", "备注"]
    output_headers = [header for header in preferred if any(clean(row.get(header)) for row in normalized_rows)]
    return {"headers": output_headers, "rows": normalized_rows}


def normalize_match_rows(source: str, rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    normalized = []
    for row in rows:
        size = clean(first_value(row, ("确认尺码", "最终尺码", "对应尺码", "自动尺码")))
        if not clean(row.get("MAKE")) and not clean(row.get("MODEL")):
            continue
        values = {key: value for key, value in row.items() if key not in LEGACY_DIMENSION_FIELDS}
        for field in DIMENSION_ALIASES:
            value = dimension_mm(row, field)
            if value:
                values[field] = value
        measurement_factor = match_measurement_factor(row)
        for field in ("T-MM", "自动长度余量", "长度余量", "相差数值"):
            if clean(row.get(field)):
                values[field] = rounded_integer(row[field], measurement_factor if field != "T-MM" else "1")
        auto_size = clean(row.get("自动尺码"))
        auto_margin = clean(row.get("自动长度余量"))
        if (
            not clean(row.get("长度余量"))
            and size
            and size not in UNAVAILABLE_SIZES
            and size == auto_size
            and auto_margin
        ):
            values["长度余量"] = rounded_integer(auto_margin, measurement_factor)
        values["CONST"] = clean(row.get("结构")) or clean(row.get("分类"))
        values["TYPE"] = clean(row.get("结构")) or clean(row.get("分类"))
        values["SIZE"] = size
        values["SOURCE"] = source
        normalized.append(
            {
                "make": clean(row.get("MAKE")),
                "model": clean(row.get("MODEL")),
                "year": clean(row.get("YEAR")),
                "years": expand_years(row.get("YEAR")),
                "construct": values["CONST"],
                "cab": clean(row.get("CAB")),
                "bed": clean(row.get("BED")),
                "type": values["TYPE"],
                "size": size,
                "values": values,
            }
        )
    return normalized


def expand_years(value: Any) -> list[int]:
    text = clean(value)
    if not text:
        return []
    parts = text.replace("–", "-").split("-")
    if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
        start, end = int(parts[0]), int(parts[1])
        if start <= end and end - start <= 150:
            return list(range(start, end + 1))
    return [int(part) for part in text.replace("/", " ").split() if part.isdigit()]


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")


def slug(value: Any, fallback: str) -> str:
    normalized = re.sub(r"[^\w.-]+", "-", clean(value).casefold(), flags=re.UNICODE).strip("-._")
    return normalized or fallback


def payload_digest(payload: Any) -> str:
    serialized = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(serialized).hexdigest()[:12]


def make_file_name(
    source_index: int,
    source_name: str,
    make_index: int,
    make_name: str,
    payload: dict[str, Any],
) -> str:
    return (
        f"size-match-{source_index:03d}-{slug(source_name, 'source')}-"
        f"{make_index:03d}-{slug(make_name, 'unknown')}-{payload_digest(payload)}.json"
    )


def write_match_payloads(path: Path, columns: list[str], sources: list[dict[str, Any]]) -> dict[str, Any]:
    """Write a small manifest plus content-addressed JSON shards per source and make."""
    path.parent.mkdir(parents=True, exist_ok=True)
    for stale_path in path.parent.glob(f"{path.stem}-*.json"):
        stale_path.unlink()

    manifest_sources: list[dict[str, Any]] = []
    for source_index, source in enumerate(sources, start=1):
        source_metadata = {
            key: value
            for key, value in source.items()
            if key != "records"
        }
        records_by_make: dict[str, tuple[str, list[dict[str, Any]]]] = {}
        for record in source.get("records") or []:
            make_name = clean(record.get("make"))
            make_key = make_name.casefold()
            if make_key not in records_by_make:
                records_by_make[make_key] = (make_name, [])
            records_by_make[make_key][1].append(record)

        make_groups: list[dict[str, Any]] = []
        ordered_groups = sorted(records_by_make.values(), key=lambda item: item[0].casefold())
        for make_index, (make_name, records) in enumerate(ordered_groups, start=1):
            shard = {
                **source_metadata,
                "make": make_name,
                "records": records,
            }
            records_path = make_file_name(
                source_index,
                clean(source.get("name")),
                make_index,
                make_name,
                shard,
            )
            write_json(path.parent / records_path, shard)
            make_groups.append(
                {
                    "make": make_name,
                    "records_path": records_path,
                    "record_count": len(records),
                }
            )

        manifest_sources.append(
            source_metadata
            | {
                "record_count": len(source.get("records") or []),
                "make_groups": make_groups,
            }
        )

    manifest = {
        "format_version": 3,
        "columns": columns,
        "sources": manifest_sources,
    }
    write_json(path, manifest)
    return manifest


def main() -> None:
    parser = argparse.ArgumentParser(description="Export frontend query JSON from an Excel workbook.")
    parser.add_argument("--xlsx-source", type=Path, help="Workbook to export. Defaults to excel_source.path in the YAML config.")
    parser.add_argument(
        "--skip-size-reference",
        action="store_true",
        help="Keep the existing size-reference JSON when the workbook only contains match sources.",
    )
    args = parser.parse_args()

    config = parse_yaml_config(CONFIG_PATH)
    source_config = config["excel_source"]
    workbook_path = args.xlsx_source.resolve() if args.xlsx_source else ROOT / source_config["path"]
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)

    if not args.skip_size_reference:
        reference_config = config["size_reference"]
        ref_headers, ref_rows = worksheet_rows(
            workbook,
            reference_config["sheet"],
            int(reference_config.get("header_row", "2")),
        )
        write_json(ROOT / reference_config["data_path"], normalize_size_reference(ref_headers, ref_rows))

    match_sources: list[dict[str, Any]] = []
    all_columns: list[str] = []
    for source in config["match_sources"]:
        headers, rows = worksheet_rows(workbook, source["sheet"], int(source.get("header_row", "1")))
        records = normalize_match_rows(source["name"], rows)
        columns = [column for column in source.get("columns", "").split(",") if column]
        if not columns:
            columns = ["MODEL", "YEAR", "TYPE", "CAB", "BED", "SIZE"]
        for column in columns:
            if column not in all_columns:
                all_columns.append(column)
        match_sources.append(
            {
                "name": source["name"],
                "label": source.get("label", source["name"]),
                "sheet": source["sheet"],
                "columns": columns,
                "records": records,
            }
        )
    write_match_payloads(ROOT / source_config["match_data_path"], all_columns, match_sources)


if __name__ == "__main__":
    main()
